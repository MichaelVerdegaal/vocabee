from openpyxl import load_workbook
import json


def get_excel_sheet():
    """
    Retrieves excel sheet from our vocabulary file. At the moment we know that we only have 1 sheet,
    so we can just retrieve the first one.
    :return: Excel sheet
    """
    workbook = load_workbook('vocab.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return sheet


def get_entries(sheet):
    """
    Retrieves vocabulary entries split by level
    :param sheet: Excel sheet from vocabulary file
    :return: Vocabulary entries as dictionary
    """
    N1 = {"entries": []}
    N2 = {"entries": []}
    N3 = {"entries": []}
    N4 = {"entries": []}
    N5 = {"entries": []}
    for row in sheet.iter_rows():
        # Each row contains a Kanji, Hiragana, English, JLPT level and Example sentences in that order
        # We put entries in a list because DataTables only accepts rows in this format
        kanji = row[0].value if row[0].value else ""
        entry = [kanji,
                 row[1].value,
                 row[2].value,
                 "Example"]

        # Split by JLPT level for performance reasons
        if row[3].value == "N1":
            N1["entries"].append(entry)
        elif row[3].value == "N2":
            N2["entries"].append(entry)
        elif row[3].value == "N3":
            N3["entries"].append(entry)
        elif row[3].value == "N4":
            N4["entries"].append(entry)
        elif row[3].value == "N5":
            N5["entries"].append(entry)

    return [N1, N2, N3, N4, N5]


def get_vocabulary():
    """
    Master function to get vocabulary and serialize it
    :return: Vocabulary entries as dictionary
    """
    sheet = get_excel_sheet()
    vocabulary = get_entries(sheet)
    for level in vocabulary:
        json.dumps(level)
    return vocabulary


vocabulary = get_vocabulary()
